#openpyxlを用いたエクセル出力
import openpyxl as op
from openpyxl.styles import Alignment as al
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabel, DataLabelList
import os.path
import math
from collections import Counter

# Check file
if(os.path.isfile("access.log") == False):
    print("access.log: No such file or directory.")   
    #exit(1)

# Read accesslog file
with open ("access.log",'r',encoding="utf8") as r_file:
    line_data = r_file.readlines()
    #close file
    r_file.close()

# Count IP function
def count_IP(line):
    IP_List = []
    for i in line:
        wordList = i.split()
        IP_List.append(wordList[0])
    l = Counter(IP_List)
    return l.most_common()

# Divide lists
length = len(line_data)
ip,access_num = zip(*count_IP(line_data))
num = len(ip)

# Edit xlsx file section
wb = op.load_workbook('sample.xlsx')
sheet = wb.worksheets[0]

# Edit value section
# Setting width
sheet.column_dimensions['A'].width = 10
sheet.column_dimensions['B'].width = 15

# Give row name
sheet['A1'].value = 'IP address'
sheet['B1'].value = 'Accessed (times)'

# Write values
for i in range(0, len(ip)):
    sheet.cell(row=i+2, column=1).value = ip[i]
    
    sheet.cell(row=i+2, column=2).value = access_num[i]

# Graph section
graph = BarChart()

graph.title = "Accessed number"
graph.x_axis.title = "IP"
graph.y_axis.title = "Accessed times"

graph.legend = None

data = Reference(sheet, min_col=1, max_col=2, min_row=1, max_row=(len(ip)+1))

cats = Reference(sheet, min_col=1, min_row=2, max_row=(len(ip)+1))

graph.add_data(data, titles_from_data=True)
graph.set_categories(cats)


sheet.add_chart(graph,'D1')



# Save xlsx file
wb.save('sample.xlsx')
# Close xlsx file
wb.close()


